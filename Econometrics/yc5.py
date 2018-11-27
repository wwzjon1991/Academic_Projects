import numpy as np
import matplotlib.pyplot as plt
import matplotlib.animation as animation
import time
import openpyxl

def update(data):
    #plt.title('Yield Curve')
    line.set_ydata(data)
    #line.set_ydata()
    time.sleep(0.1)
    return line


wb = openpyxl.load_workbook('DTYCR.xlsx')
sheet = wb.get_sheet_by_name('DTYCR')
rows = sheet.max_row
cols = sheet.max_column

DATES = []
for rowOfCellObjects in sheet['A3001':'A5000']:
    for obj in rowOfCellObjects:
        d = str(obj.value)
        DATES.append(d[0:10])
     
nd = len(DATES)
for i in range(nd):
    if DATES[i] == '2005-07-01':
        break
si = i
for i in range(nd):
    if DATES[i] == '2009-09-30':
        break
ei = i
    
   
def Read_Column(n, sr, er):
    C = np.zeros(n, dtype='float')     
    i = 0
    for rowOfCellObjects in sheet[sr:er]:
        for cellObj in rowOfCellObjects:
            tcov = type(cellObj.value)
            if tcov is unicode or tcov is str:
                v = float('nan')
            else:
                v = cellObj.value            
            C[i] = v
            i+=1
    return C
     
M1 = Read_Column(2000,'B3001', 'B5000')
M3 = Read_Column(2000,'C3001', 'C5000')
M6 = Read_Column(2000,'D3001', 'D5000')
Y1 = Read_Column(2000,'E3001', 'E5000')
Y2 = Read_Column(2000,'F3001', 'F5000')
Y3 = Read_Column(2000,'G3001', 'G5000')
Y5 = Read_Column(2000,'H3001', 'H5000')
Y7 = Read_Column(2000,'I3001', 'I5000')
Y10 = Read_Column(2000,'J3001', 'J5000')

fig, ax = plt.subplots()
mat = [1/12.0, 1/4.0, 1/2.0, 1, 2, 3, 5, 7, 10]
line, = ax.plot(mat, np.zeros(len(mat), dtype='float'), '-bo')
ax.set_ylim(0, 6)
ax.grid()

pause = False
def data_gen(i,j,m1,m3,m6,y1,y2,y3,y5,y7,y10,dates):
    while (i < j): 
        if not pause:
            y = [m1[i], m3[i], m6[i], y1[i], y2[i], y3[i], y5[i], y7[i], y10[i]]
            y = np.array(y)
            plt.title(dates[i])
        yield y
        i += 1

def onClick(event):
    global pause
    pause ^= True

fig.canvas.mpl_connect('button_press_event', onClick)        
ani = animation.FuncAnimation(fig, update, data_gen(si, ei, M1, M3, M6, Y1, Y2, Y3, Y5, Y7, Y10, DATES), 
                              interval=300, repeat=False) 
# interval=500)# repeat_delay=5000)
plt.show()

